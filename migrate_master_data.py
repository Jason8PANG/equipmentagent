"""
将 设备维修数据统计表.xlsx 中的主数据迁移到 MySQL
- 设备清单 sheet → 设备清单 表
- 机器故障类别 sheet → 故障类别明细 表
- 技术员 sheet → 维修技术员 表
"""
import pymysql
import openpyxl

MYSQL_CONFIG = {
    'host': '10.0.6.86',
    'port': 33306,
    'user': 'powerbi',
    'password': '!Q1234567',
    'database': 'erp_data',
    'charset': 'utf8mb4',
}

EXCEL_PATH = r'D:\02 Quality-Yan\03 Plant2\04 Copilot\2.设备维修数据\设备维修数据库\设备维修数据统计表.xlsx'

conn = pymysql.connect(**MYSQL_CONFIG)
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

# ══════════════════════════════════════════════
# 1. 设备清单
# ══════════════════════════════════════════════
print("=== 1. 设备清单 ===")
ws = wb['设备清单']

with conn.cursor() as c:
    c.execute("""
        CREATE TABLE IF NOT EXISTS `设备清单` (
            `id` INT AUTO_INCREMENT PRIMARY KEY,
            `设备名称` VARCHAR(100) NOT NULL,
            `设备型号` VARCHAR(100),
            `设备编号` VARCHAR(100) NOT NULL UNIQUE,
            `设备数量` INT DEFAULT 1
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    c.execute("DELETE FROM `设备清单`")
conn.commit()

equipments = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    name = str(row[0]).strip() if row[0] else ''
    model = str(row[1]).strip() if row[1] else ''
    code = str(row[2]).strip() if row[2] else ''
    qty  = int(row[3]) if row[3] else 1
    if name and code:
        equipments.append((name, model, code, qty))
        print(f"  {name} | {model} | {code} | {qty}")

with conn.cursor() as c:
    c.executemany(
        "INSERT INTO `设备清单` (`设备名称`, `设备型号`, `设备编号`, `设备数量`) VALUES (%s, %s, %s, %s)",
        equipments
    )
conn.commit()
print(f"  共 {len(equipments)} 条\n")

# ══════════════════════════════════════════════
# 2. 故障类别明细
# ══════════════════════════════════════════════
print("=== 2. 故障类别明细 ===")
ws = wb['机器故障类别']

with conn.cursor() as c:
    c.execute("""
        CREATE TABLE IF NOT EXISTS `故障类别明细` (
            `id` INT AUTO_INCREMENT PRIMARY KEY,
            `故障类别` VARCHAR(50) NOT NULL,
            `故障明细` VARCHAR(100) NOT NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    c.execute("DELETE FROM `故障类别明细`")
conn.commit()

current_cat = None
fault_details = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    cat = str(row[0]).strip() if row[0] else ''
    detail = str(row[1]).strip() if row[1] else ''
    if cat:
        current_cat = cat
    if current_cat and detail:
        fault_details.append((current_cat, detail))
        print(f"  [{current_cat}] {detail}")

with conn.cursor() as c:
    c.executemany(
        "INSERT INTO `故障类别明细` (`故障类别`, `故障明细`) VALUES (%s, %s)",
        fault_details
    )
conn.commit()
print(f"  共 {len(fault_details)} 条\n")

# ══════════════════════════════════════════════
# 3. 维修技术员
# ══════════════════════════════════════════════
print("=== 3. 维修技术员 ===")
ws = wb['技术员']

with conn.cursor() as c:
    c.execute("""
        CREATE TABLE IF NOT EXISTS `维修技术员` (
            `id` INT AUTO_INCREMENT PRIMARY KEY,
            `工号` VARCHAR(20) NOT NULL UNIQUE,
            `姓名` VARCHAR(50) NOT NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    c.execute("DELETE FROM `维修技术员`")
conn.commit()

techs = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    emp_id = str(row[0]).strip() if row[0] else ''
    name = str(row[1]).strip() if row[1] else ''
    if emp_id and name:
        techs.append((emp_id, name))
        print(f"  {emp_id} | {name}")

with conn.cursor() as c:
    c.executemany(
        "INSERT INTO `维修技术员` (`工号`, `姓名`) VALUES (%s, %s)",
        techs
    )
conn.commit()
print(f"  共 {len(techs)} 条\n")

# ══════════════════════════════════════════════
# 4. 验证
# ══════════════════════════════════════════════
print("=== 验证 ===")
with conn.cursor() as c:
    c.execute("SELECT COUNT(*) FROM `设备清单`")
    print(f"  设备清单: {c.fetchone()[0]} 条")
    c.execute("SELECT COUNT(*) FROM `故障类别明细`")
    print(f"  故障类别明细: {c.fetchone()[0]} 条")
    c.execute("SELECT COUNT(*) FROM `维修技术员`")
    print(f"  维修技术员: {c.fetchone()[0]} 条")
    c.execute("SELECT DISTINCT `故障类别` FROM `故障类别明细`")
    cats = [row[0] for row in c.fetchall()]
    print(f"  故障类别: {cats}")
    for cat in cats:
        c.execute("SELECT `故障明细` FROM `故障类别明细` WHERE `故障类别` = %s ORDER BY `故障明细`", (cat,))
        details = [row[0] for row in c.fetchall()]
        print(f"    [{cat}]: {details}")

conn.close()
wb.close()
print("\nDone!")
